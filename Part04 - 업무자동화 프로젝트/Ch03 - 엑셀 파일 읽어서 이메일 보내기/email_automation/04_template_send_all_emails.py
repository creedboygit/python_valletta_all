import os
import smtplib
from email.mime.text import MIMEText

from openpyxl import load_workbook


class EmailSender:
    email_addr = None
    password = None
    smtp_server_map = {
        "gmail.com": "smtp.gmail.com",
        "naver.com": "smtp.naver.com"
    }
    smtp_server = None

    def __init__(self, email_addr, password):
        print("생성자")
        self.email_addr = email_addr
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]]
        # print(self.smtp_server)

    def send_email(self, msg, from_addr, to_addr, subject):
        """
        :param subject: 제목
        :param msg: 보낼 메시지
        :param from_addr: 보내는 사람
        :param to_addr: 받는 사람
        """

        # with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEText(msg)
            msg['From'] = from_addr
            msg['To'] = to_addr
            msg['Subject'] = subject
            # print(msg.as_string())

            smtp.starttls()
            # smtp.login(self.email_addr, self.password)
            smtp.login("polozh@naver.com", self.password)
            # smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.encode("utf-8"))
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print(f"{to_addr} 로 이메일 전송이 완료되었습니다.")

    def send_all_emails(self, filename):
        print(f"{filename}에 있는 이메일과 내용을 이용하여 이메일을 보냅니다.")
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            temp1 = """
안녕하세요. 네이버 쇼핑몰입니다.

            
"""
            if row[0].value is not None:
                print(row[0].value, row[1].value, row[2].value)
                self.send_email(msg=row[2].value,
                                from_addr=self.email_addr,
                                to_addr=row[0].value,
                                subject=row[1].value)


if __name__ == '__main__':
    # es = EmailSender("polozhzh@gmail.com", os.getenv("MY_GMAIL_PASSWORD"))
    # es.send_email("테스트입니다 2", from_addr="polozhzh@gmail.com", to_addr="polozh@naver.com")
    es = EmailSender("polozh@naver.com", os.getenv("MY_NAVER_PASSWORD"))
    # es.send_email("테스트입니다 2", from_addr="polozh@naver.com", to_addr="polozhzh@gmail.com")
    # es.send_email("테스트입니다 2",
    #               from_addr="polozh@naver.com",
    #               to_addr="polozhzh@gmail.com",
    #               subject="이메일 발송 테스트입니다.")
    es.send_all_emails("sample.xlsx")
