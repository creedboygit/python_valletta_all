import os
from datetime import datetime

import pandas as pd
from dateutil.utils import today
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.utils import get_column_letter

pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)


class ClassificatinoExcel:
    path = ""
    now_day = ""

    def __init__(self, order_xlsx_filename, partner_info_xlsx_filename, path="result"):
        # 주문 목록
        df = pd.read_excel(order_xlsx_filename)
        df = df.rename(columns=df.iloc[1])  # 열 제목으로 1번째를 지정
        df = df.drop([df.index[0], df.index[1]])
        # print("======= df.count():\n" + str(df.count()))
        # print(df['상품명'])

        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path
        self.now_day = datetime.now().strftime("%Y-%m-%d")

        if not os.path.exists(self.path):
            os.mkdir(self.path)

        # print(df)

        # 파트너 목록
        df_partners = pd.read_excel(partner_info_xlsx_filename)
        # print(df_partners)
        self.brands = df_partners['브랜드'].to_list()
        self.partners = df_partners['업체명'].to_list()

        # print("======= self.brands:\n" + str(self.brands), len(self.brands))
        # print("======= self.partners:\n" + str(self.partners), len(self.partners))
        #
        # print(self.brands[0], self.partners[0])
        # print(self.order_list.head())
        # print(self.order_list['상품명'].head())

    def classify(self):
        # for i, row, in self.order_list.head(5).iterrows():
        for i, row, in self.order_list.iterrows():
            brand_name = ""
            partner_name = ""
            for j in range(len(self.brands)):
                # print(self.brands[j])
                if self.brands[j] in row['상품명']:
                    # print(f"{self.brands[j]}이(가) {j}번째에 포함되어 있습니다.")
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break
            # print(f"{row['상품명']} 은 {brand_name} 브랜드입니다. {j}번째")
            # print(f"업체명: {partner_name}")

            if partner_name != "":
                # print(self.order_list['상품명'].str.contains(brand_name), brand_name)
                # asis = self.order_list['상품명'].str.contains(brand_name)
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                df_filtered.to_excel(f"{self.path}/[쇼핑몰] {partner_name}.xlsx", index=False)
            else:
                print("없는 브랜드명입니다: ", brand_name, row["상품명"])

    def set_form(self, filename):
        # filename = "2024-03-02/[쇼핑몰] 더블유데이.xlsx"
        wb = load_workbook(filename)
        ws = wb.active
        # print("value: ", ws['B1'].value)

        # 개수 세기
        row_cnt = ws.max_row - 1
        print(f"{filename} cnt: ", row_cnt)

        # 열 삽입
        ws.insert_rows(1)
        ws.insert_rows(1)

        # A1
        # now_day = datetime.now().strftime("%Y-%m-%d")
        ws['A1'] = f"발송요청내역 [총 {row_cnt}건] - {self.now_day}"
        ws['A1'].font = Font(size=11, bold=True)
        ws.merge_cells("a1:u1")
        ws['a1'].alignment = Alignment(horizontal='left')

        # 열 너비
        # 1 ~ 25(Y)
        for i in range(1, 25 + 1):
            ws.column_dimensions[get_column_letter(i)].width = 13

        for col_letter in ['I', 'J', 'W', 'X']:
            ws.column_dimensions[col_letter].width = 40

        for col_letter in ['A', 'V', 'H']:
            ws.column_dimensions[col_letter].width = 26

        # 컬럼명에 색상, bold
        for row in ws.iter_rows(min_row=3, max_row=3):
            for cell in row:
                cell.fill = PatternFill(fgColor='cccccc', fill_type='solid')

        # 주문 목록
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', shrink_to_fit=True)
                cell.fill = PatternFill(fgColor='ffffcc', fill_type='solid')
                cell.border = Border(left=Side(border_style=BORDER_THIN),
                                     right=Side(border_style=BORDER_THIN),
                                     top=Side(border_style=BORDER_THIN),
                                     bottom=Side(border_style=BORDER_THIN))

        wb.save(filename)

    def set_forms(self):
        file_list = os.listdir(self.path)
        # print(file_list)

        for filename in file_list:
            file_name = f"{self.path}/{filename}"
            self.set_form(file_name)


if __name__ == '__main__':
    # print("======= today:\n" + str(today()))

    # now_day = datetime.now().strftime("%Y-%m-%d")

    ce = ClassificatinoExcel("주문목록20221112.xlsx", "파트너목록.xlsx", str(today()).replace(":", "")[:10])
    ce.classify()
    ce.set_forms()
