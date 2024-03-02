import os
from datetime import datetime

import pandas as pd
from dateutil.utils import today
from openpyxl import load_workbook

pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)


class ClassificatinoExcel:
    path = ""

    def __init__(self, order_xlsx_filename, partner_info_xlsx_filename, path="result"):
        # 주문 목록
        df = pd.read_excel(order_xlsx_filename)
        df = df.rename(columns=df.iloc[1])  # 열 제목으로 1번째를 지정
        df = df.drop([df.index[0], df.index[1]])
        # print("======= df.count():\n" + str(df.count()))
        # print(df['상품명'])

        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path

        if not os.path.exists(self.path):
            os.mkdir(self.path)

        # print(df)

        # 파트너 목록
        df_partners = pd.read_excel(partner_info_xlsx_filename)
        # print(df_partners)
        self.brands = df_partners['브랜드'].to_list()
        self.partners = df_partners['업체명'].to_list()

        # print("======= self.brands:\n" + str(self.brands), len(self.brands))
        # print("======= self.partners:\n" + str(self.partners), len(self.partners))
        #
        # print(self.brands[0], self.partners[0])
        # print(self.order_list.head())
        # print(self.order_list['상품명'].head())

    def classify(self):
        # for i, row, in self.order_list.head(5).iterrows():
        for i, row, in self.order_list.iterrows():
            brand_name = ""
            partner_name = ""
            for j in range(len(self.brands)):
                # print(self.brands[j])
                if self.brands[j] in row['상품명']:
                    # print(f"{self.brands[j]}이(가) {j}번째에 포함되어 있습니다.")
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break
            # print(f"{row['상품명']} 은 {brand_name} 브랜드입니다. {j}번째")
            # print(f"업체명: {partner_name}")

            if partner_name != "":
                # print(self.order_list['상품명'].str.contains(brand_name), brand_name)
                # asis = self.order_list['상품명'].str.contains(brand_name)
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                df_filtered.to_excel(f"{self.path}/[쇼핑몰] {partner_name}.xlsx")
            else:
                print("없는 브랜드명입니다: ", brand_name, row["상품명"])

    def set_count(self):
        filename = "2024-03-02/[쇼핑몰] 더블유데이.xlsx"
        wb = load_workbook(filename)
        ws = wb.active
        # print("value: ", ws['B1'].value)

        # 개수 세기
        row_cnt = ws.max_row - 1
        # print("cnt: ", row_cnt)

        # 열 삽입
        ws.insert_rows(1)
        ws.insert_rows(1)

        # A1
        now_day = datetime.now().strftime("%Y-%m-%d")
        ws['A1'] = f"발송요청내역 [총 {row_cnt}건] - {now_day}"

        wb.save(filename)


if __name__ == '__main__':
    # print("======= today:\n" + str(today()))

    now_day = datetime.now().strftime("%Y-%m-%d")

    ce = ClassificatinoExcel("주문목록20221112.xlsx", "파트너목록.xlsx", str(today()).replace(":", "")[:10])
    # ce.classify()
    ce.set_count()
