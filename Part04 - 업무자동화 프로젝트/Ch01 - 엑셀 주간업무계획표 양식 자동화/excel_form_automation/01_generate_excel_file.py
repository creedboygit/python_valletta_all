from openpyxl import Workbook


class WeeklyWorkPlan:
    wb = None
    ws = None
    manager = "매니저 이름을 입력해 주세요."
    start_date = "2024-03-11"

    def __init__(self, manager, start_date, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.manager = manager
        self.start_date = start_date
        self.set_title()

    def save(self, filename):
        self.wb.save(filename)
        print("엑셀 파일 생성 완료")

    def set_title(self):
        self.ws['B2'] = "담당자"
        self.ws['C2'] = self.manager


if __name__ == '__main__':
    wwp = WeeklyWorkPlan("크리드")
    wwp.save("주간업무계획표.xlsx")
