from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.utils import get_column_letter


class WeeklyWorkPlan:
    wb = None
    ws = None
    manager = "매니저 이름을 입력해 주세요."
    start_date = "2024-03-11"
    date_list = []
    days_of_week = []
    weekday_list = ['월', '화', '수', '목', '금', '토', '일']

    def __init__(self, manager, start_date, days=5, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.manager = manager
        self.start_date = start_date

        # 날짜 생성
        # self.set_date()
        # self.set_date(5)
        self.set_date(days=days)
        self.set_title()
        self.set_table()
        self.set_style()

    def save(self, filename):
        self.wb.save(filename)
        print("엑셀 파일 생성 완료")

    def set_title(self):
        ws = self.ws
        ws['B2'] = "담당자"
        ws['C2'] = self.manager

        ws['B3'] = "시작일"
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = "주간업무계획표"

        # 날짜
        # start_date = "2024-03-11"
        # end_date = "2024-03-17"
        start_date = self.date_list[0]
        end_date = self.date_list[-1]
        ws['B6'] = f"({start_date} ~ {end_date})"

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print("타이틀 생성 완료")

    def set_table(self):
        ws = self.ws
        ws['B8'] = "날짜"
        col_names = ['날짜', '요일', '시간', '일정', '비고']

        # 컬럼명
        for i in range(5):
            ws.cell(row=8, column=(i + 2)).value = col_names[i]

        # 날짜, 요일
        for i in range(len(self.date_list)):
            # ws.cell(row=9 + i, column=2).value = self.date_list[i]
            ws.cell(row=9 + (i * 5), column=2).value = self.date_list[i]  # 날짜
            ws.cell(row=9 + (i * 5), column=3).value = self.days_of_week[i]  # 요일

            # 셀 병합
            ws.merge_cells(f'B{i * 5 + 9}:B{i * 5 + 13}')  # 날짜
            ws.merge_cells(f'C{i * 5 + 9}:C{i * 5 + 13}')  # 요일
            ws.merge_cells(f'F{i * 5 + 9}:F{i * 5 + 13}')  # 비고

    def set_date(self, days=6):
        # start_date + 6일
        end_date = datetime.strptime(self.start_date, "%Y-%m-%d") + timedelta(days=days)
        # [5, 6, 7, 8, 9, 10, 11]
        week = pd.date_range(start=self.start_date,
                             end=end_date.strftime("%Y-%m-%d"))
        self.date_list = week.strftime("%Y-%m-%d").to_list()
        self.days_of_week = week.strftime("%A").to_list()
        # apply = pd.DataFrame.apply(lambda x: self.weekday_list[x['self.days_of_week']], axis=1)
        print(end_date)
        print("======= week:\n" + str(week))
        print("======= date_list:\n" + str(self.date_list))
        print("======= days_of_week:\n" + str(self.days_of_week))

    def set_style(self):
        ws = self.ws

        # A열 너비
        ws.column_dimensions['A'].width = 5

        # 컬럼명
        for i in range(2, 7):
            # print(get_column_letter(i))
            # print("======= get_column_letter(i):\n" + str(i) + " - " + str(get_column_letter(i)))
            ws.column_dimensions[get_column_letter(i)].width = 15
            ws[f'{get_column_letter(i)}8'].font = Font(name="맑은 고딕", bold=True)
            ws[f'{get_column_letter(i)}8'].alignment = Alignment(horizontal='center', vertical='center')

            # fill 색상 채우기
            ws[f'{get_column_letter(i)}8'].fill = PatternFill(fgColor='cccccc', fill_type='solid')

        # E열 너비
        ws.column_dimensions['E'].width = 40

        # 제목 폰트, 사이즈
        ws['B5'].font = Font(name="맑은 고딕", size=28, bold=True)

        # 가운데 정렬
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

        # 날짜, 요일 가운데 정렬
        for i in range(9, 40, 5):
            ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{i}'].alignment = Alignment(horizontal='center', vertical='center')

        # 담당자, 시작일 fill color
        ws['B2'].fill = PatternFill(fgColor='cccccc', fill_type='solid')
        ws['B3'].fill = PatternFill(fgColor='cccccc', fill_type='solid')

        # 테두리 설정 - border
        border_style = Border(left=Side(border_style=BORDER_THIN),
                              right=Side(border_style=BORDER_THIN),
                              top=Side(border_style=BORDER_THIN),
                              bottom=Side(border_style=BORDER_THIN))

        # 담당자, 시작일 border
        ws['B2'].border = border_style
        ws['C2'].border = border_style
        ws['B3'].border = border_style
        ws['C3'].border = border_style

        # 표 영역 border - 13, 18, 23 ... 38
        # for col in ws.iter_cols(min_row=8, min_col=2, max_row=38, max_col=6):
        for col in ws.iter_cols(min_row=8, min_col=2, max_row=len(self.date_list) * 5 + 8, max_col=6):
            # print("======= col:\n" + str(col))

            for cell in col:
                cell.border = border_style


if __name__ == '__main__':
    wwp = WeeklyWorkPlan("크리드", "2024-03-11", days=6)
    wwp.save("주간업무계획표.xlsx")
