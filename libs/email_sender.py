import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr

from openpyxl import load_workbook


class EmailSender:
    email_addr = None
    manager_name = None
    password = None
    smtp_server_map = {
        "gmail.com": "smtp.gmail.com",
        "naver.com": "smtp.naver.com"
    }
    smtp_server = None
    template_file_name = None

    def __init__(self, manager_name, email_addr, password, template_file_name):
        print("생성자")
        self.email_addr = email_addr
        self.manager_name = manager_name
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]]
        # print(self.smtp_server)
        self.template_file_name = template_file_name

    def send_email(self, html_msg, from_addr, to_addr, to_name, subject):

        # with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            # msg = MIMEText(msg)
            msg = MIMEMultipart("alternative")
            # msg['From'] = from_addr
            # msg['From'] = formataddr(("무역회사", from_addr))
            msg['From'] = formataddr((self.manager_name, from_addr))
            msg['To'] = formataddr((to_name, to_addr))
            msg['Subject'] = subject + str(datetime.now())
            msg.attach(MIMEText(html_msg, 'html', 'utf-8'))
            # print(msg.as_string())

            smtp.starttls()
            # smtp.login(self.email_addr, self.password)
            smtp.login("polozh@naver.com", self.password)
            # smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.encode("utf-8"))
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print(f"{to_addr} 로 이메일 전송이 완료되었습니다.")

    def send_all_emails(self, filename):
        print(f"{filename}에 있는 이메일과 내용을 이용하여 이메일을 보냅니다.")
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None:
                with open(self.template_file_name, encoding='utf-8') as f:
                    temp1 = f.read()

                    print("======= temp1:\n" + str(temp1))

                    print(row[0].value, row[1].value, row[2].value)
                    temp1 = (temp1.replace("%받는분%", row[1].value)
                             .replace("%매니저명%", self.manager_name))
                    self.send_email(html_msg=temp1,
                                    from_addr=self.email_addr,
                                    to_addr=row[0].value,
                                    to_name=row[1].value,
                                    subject=row[2].value)
            else:
                print('row[0]이 None입니다.')


if __name__ == '__main__':
    # es = EmailSender("polozhzh@gmail.com", os.getenv("MY_GMAIL_PASSWORD"))
    # es.send_email("테스트입니다 2", from_addr="polozhzh@gmail.com", to_addr="polozh@naver.com")
    es = EmailSender(manager_name="보이스", email_addr="polozh@naver.com", password=os.getenv("MY_NAVER_PASSWORD"))
    # es.send_email("테스트입니다 2", from_addr="polozh@naver.com", to_addr="polozhzh@gmail.com")
    # es.send_email("테스트입니다 2",
    #               from_addr="polozh@naver.com",
    #               to_addr="polozhzh@gmail.com",
    #               subject="이메일 발송 테스트입니다.")
    es.send_all_emails("sample_with_name.xlsx")
